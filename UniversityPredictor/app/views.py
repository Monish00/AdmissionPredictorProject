from django.shortcuts import render,redirect
from django.contrib import messages
from . models import *
import random 
from django.db.models import Sum, Count
from django.db import connection
from django.core.mail import send_mail
from django.core.mail import EmailMessage
from django.conf import settings
from datetime import datetime, timedelta
from datetime import date 
import datetime
import openpyxl

def student_login(request):
	if request.session.has_key('user_id'):
		return render(request,'dashboard.html',{})
	else:
		if request.method == 'POST':
			name=request.POST.get('username')
			pwd=request.POST.get('password')
			user_exist=Student_Detail.objects.filter(username=name,password=pwd)
			if user_exist:
				request.session['name']= request.POST.get('username')
				a = request.session['name']
				sess = Student_Detail.objects.only('id').get(username=a).id
				request.session['user_id']= sess
				return redirect('dashboard')
			else:
				messages.success(request,'Invalid username or Password')
		return render(request,'student_login.html',{})
def dashboard(request):
	return render(request,'dashboard.html',{})
def register(request):
	if request.method == 'POST':
		Name = request.POST.get('uname')
		Adddress = request.POST.get('address')
		Mobile= request.POST.get('mobile')
		Email = request.POST.get('email')
		Password = request.POST.get('pwd')
		unum = request.POST.get('username')
		country = request.POST.get('country')
		city = request.POST.get('city')
		state = request.POST.get('state')
		dob = request.POST.get('dob')
		gender = request.POST.get('gender')
		education = request.POST.get('education')
		cutoff_mark = request.POST.get('cutoff_mark')
		score = request.POST.get('score')
		image = request.FILES['image']
		caste =request.POST.get('caste')
		student_exist = Student_Detail.objects.filter(username=unum)
		if student_exist:
			messages.success(request,'Username No Already Exsit')
		else:
			crt = Student_Detail.objects.create(student_name=Name,
			address=Adddress,phone_number=Mobile,password=Password,email_id=Email,username=unum,country=country,
			city=city,state=state,dob=dob,caste=caste,gender=gender,education=education,Score=score,image=image,cutoff_mark=cutoff_mark)
			if crt:
				messages.success(request,'Registered Successfully')
	return render(request,'register.html',{})
def logout(request):
    try:
        del request.session['user_id']
        del request.session['name']
    except:
     pass
    return render(request, 'student_login.html', {})
def admin_home(request):
	if request.method == 'POST':
		file = request.FILES['excel_file']	  
		workbook = openpyxl.load_workbook(file, read_only=True)
		# Get name of the first sheet and then open sheet by name
		first_sheet = workbook.get_sheet_names()[0]
		worksheet = workbook.get_sheet_by_name(first_sheet)
		data = []
		for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row): 
			detail = Engineering_College_Detail()
			
			detail.college_name = row[0].value
			detail.degree = row[1].value
			detail.department = row[2].value
			detail.OC = row[3].value
			detail.BC = row[4].value
			detail.BCM = row[5].value
			detail.MBCV = row[6].value
			detail.MBCDNC = row[7].value
			detail.MBA = row[8].value
			detail.SC = row[9].value
			detail.SCA = row[10].value
			detail.ST = row[11].value
			data.append(detail)
		# Bulk create data
		crt = Engineering_College_Detail.objects.bulk_create(data)
		if crt:
			messages.success(request,"Detail Added Successfully.")
	return render(request,'home.html',{})
def agri_search(request):
	if request.method == 'POST':
		file = request.FILES['excel_file']	  
		workbook = openpyxl.load_workbook(file, read_only=True)
		# Get name of the first sheet and then open sheet by name
		first_sheet = workbook.get_sheet_names()[0]
		worksheet = workbook.get_sheet_by_name(first_sheet)
		data = []
		for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row): 
			detail = Agri_College_Detail()
			
			detail.college_name = row[0].value
			detail.degree = row[1].value
			detail.department = row[2].value
			detail.OC = row[3].value
			detail.BC = row[4].value
			detail.BCM = row[5].value
			detail.MBCV = row[6].value
			detail.MBCDNC = row[7].value
			detail.MBA = row[8].value
			detail.SC = row[9].value
			detail.SCA = row[10].value
			detail.ST = row[11].value
			data.append(detail)
		# Bulk create data
		crt = Agri_College_Detail.objects.bulk_create(data)
		if crt:
			messages.success(request,"Detail Added Successfully.")
	return render(request,'agri_search.html',{})
def medical(request):
	if request.method == 'POST':
		file = request.FILES['excel_file']	  
		workbook = openpyxl.load_workbook(file, read_only=True)
		# Get name of the first sheet and then open sheet by name
		first_sheet = workbook.get_sheet_names()[0]
		worksheet = workbook.get_sheet_by_name(first_sheet)
		data = []
		for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row): 
			detail = Medical_College_Detail()
			
			detail.college_name = row[0].value
			detail.degree = row[1].value
			detail.department = row[2].value
			detail.OC = row[3].value
			detail.BC = row[4].value
			detail.BCM = row[5].value
			detail.MBCV = row[6].value
			detail.MBCDNC = row[7].value
			detail.MBA = row[8].value
			detail.SC = row[9].value
			detail.SCA = row[10].value
			detail.ST = row[11].value
			data.append(detail)
		# Bulk create data
		crt = Medical_College_Detail.objects.bulk_create(data)
		if crt:
			messages.success(request,"Detail Added Successfully.")
	return render(request,'medical.html',{})
def engineering_search(request):
	user_id=request.session['user_id']
	cur=connection.cursor()
	user = '''SELECT s.caste from app_student_detail as s where s.id='%d' ''' %(int(user_id))
	sel = cur.execute(user)
	user_detail = cur.fetchone()
	caste = user_detail[0]
	cursor=connection.cursor()
	sql=''' SELECT e.degree from app_engineering_college_detail as e GROUP BY e.degree'''
	post=cursor.execute(sql)
	row=cursor.fetchall()
	cursor1=connection.cursor()
	sql1=''' SELECT e.department from app_engineering_college_detail as e GROUP BY e.department'''
	post1=cursor1.execute(sql1)
	row1=cursor1.fetchall()
	if request.method == 'POST':
		degree=request.POST.get('degree')
		department=request.POST.get('department')
		cutoff_mark =request.POST.get('cutoff_mark')
		if caste == 'OC':
			a=Engineering_College_Detail.objects.filter(degree=degree,department=department,OC__lte=cutoff_mark)
			cursor=connection.cursor()
			sql = '''SELECT e.college_name,e.degree,e.department from app_engineering_college_detail as e where e.degree='%s' 
			AND e.department='%s' AND e.OC='%s' ''' % (degree,department,cutoff_mark)
			post=cursor.execute(sql)
			col_detail = cursor.fetchone()
			if col_detail:
				col_name = col_detail[0]
				import requests
				import urllib
				import pandas as pd
				from requests_html import HTML
				from requests_html import HTMLSession
				def get_source(url):
				    """Return the source code for the provided URL. 

				    Args: 
				        url (string): URL of the page to scrape.

				    Returns:
				        response (object): HTTP response object from requests_html. 
				    """

				    try:
				        session = HTMLSession()
				        response = session.get(url)
				        return response

				    except requests.exceptions.RequestException as e:
				        print(e)

				def scrape_google(query):

					query = urllib.parse.quote_plus(query)
					response = get_source("https://www.google.com/search?q=" + query)

					links = list(response.html.absolute_links)
					google_domains = ('https://www.google.'
					)

					for url in links[:]:
						if url.startswith(google_domains):
							links.remove(url)

					return links

				link=scrape_google(col_name)
				

				return render(request,'engineering_search.html',{'col_detail':col_detail,'row':row,'row1':row1,'a':a,'link':link})
		elif caste == 'BC':
			a=Engineering_College_Detail.objects.filter(degree=degree,department=department,BC__lte=cutoff_mark)
			cursor=connection.cursor()
			sql1 = '''SELECT e.college_name,e.degree,e.department,e.BC from app_engineering_college_detail as e where e.degree='%s' 
			AND e.department='%s' AND e.BC <= '%s' ''' % (degree,department,cutoff_mark)
			post1=cursor.execute(sql1)
			row_1 = cursor.fetchall()
			cursor=connection.cursor()
			sql = '''SELECT e.college_name,e.degree,e.department,e.BC from app_engineering_college_detail as e where e.degree='%s' 
			AND e.department='%s' AND e.BC <= '%s' ''' % (degree,department,cutoff_mark)
			post=cursor.execute(sql)
			col_detail = cursor.fetchone()
			if col_detail:
				col_name = col_detail[0]
				import requests
				import urllib
				import pandas as pd
				from requests_html import HTML
				from requests_html import HTMLSession
				def get_source(url):
				    """Return the source code for the provided URL. 

				    Args: 
				        url (string): URL of the page to scrape.

				    Returns:
				        response (object): HTTP response object from requests_html. 
				    """

				    try:
				        session = HTMLSession()
				        response = session.get(url)
				        return response

				    except requests.exceptions.RequestException as e:
				        print(e)

				def scrape_google(query):

					query = urllib.parse.quote_plus(query)
					response = get_source("https://www.google.com/search?q=" + query)

					links = list(response.html.absolute_links)
					google_domains = ('https://www.google.'
					)

					for url in links[:]:
						if url.startswith(google_domains):
							links.remove(url)

					return links

				link=scrape_google(col_name)
				return render(request,'engineering_search.html',{'link':link,'row':row,
				'row1':row1,'a':a,'col_detail':col_detail,'row_1':row_1})
		elif caste == 'BCM':
			a=Engineering_College_Detail.objects.filter(degree=degree,department=department,BCM__lte=cutoff_mark)
			cursor=connection.cursor()
			sql1 = '''SELECT e.college_name,e.degree,e.department,e.BCM from app_engineering_college_detail as e where e.degree='%s' 
			AND e.department='%s' AND e.BCM <= '%s' ''' % (degree,department,cutoff_mark)
			post1=cursor.execute(sql1)
			row_1 = cursor.fetchall()
			cursor=connection.cursor()
			sql = '''SELECT e.college_name,e.degree,e.department,e.BCM from app_engineering_college_detail as e where e.degree='%s' 
			AND e.department='%s' AND e.BCM='%s' ''' % (degree,department,cutoff_mark)
			post=cursor.execute(sql)
			col_detail = cursor.fetchone()
			if col_detail:
				col_name = col_detail[0]
				import requests
				import urllib
				import pandas as pd
				from requests_html import HTML
				from requests_html import HTMLSession
				def get_source(url):
				    """Return the source code for the provided URL. 

				    Args: 
				        url (string): URL of the page to scrape.

				    Returns:
				        response (object): HTTP response object from requests_html. 
				    """

				    try:
				        session = HTMLSession()
				        response = session.get(url)
				        return response

				    except requests.exceptions.RequestException as e:
				        print(e)

				def scrape_google(query):

					query = urllib.parse.quote_plus(query)
					response = get_source("https://www.google.com/search?q=" + query)

					links = list(response.html.absolute_links)
					google_domains = ('https://www.google.'
					)

					for url in links[:]:
						if url.startswith(google_domains):
							links.remove(url)

					return links

				link=scrape_google(col_name)
				return render(request,'engineering_search.html',{'link':link,'row':row,
				'row1':row1,'a':a,'col_detail':col_detail,'row_1':row_1})
		elif caste == 'MBCV':
			a=Engineering_College_Detail.objects.filter(degree=degree,department=department,MBCV__lte=cutoff_mark)
			cursor=connection.cursor()
			sql1 = '''SELECT e.college_name,e.degree,e.department,e.MBCV from app_engineering_college_detail as e where e.degree='%s' 
			AND e.department='%s' AND e.MBCV <= '%s' ''' % (degree,department,cutoff_mark)
			post1=cursor.execute(sql1)
			row_1 = cursor.fetchall()
			cursor=connection.cursor()
			sql = '''SELECT e.college_name,e.degree,e.department from app_engineering_college_detail as e where e.degree='%s' 
			AND e.department='%s' AND e.MBCV='%s' ''' % (degree,department,cutoff_mark)
			post=cursor.execute(sql)
			col_detail = cursor.fetchone()
			if col_detail:
				col_name = col_detail[0]
				import requests
				import urllib
				import pandas as pd
				from requests_html import HTML
				from requests_html import HTMLSession
				def get_source(url):
				    """Return the source code for the provided URL. 

				    Args: 
				        url (string): URL of the page to scrape.

				    Returns:
				        response (object): HTTP response object from requests_html. 
				    """

				    try:
				        session = HTMLSession()
				        response = session.get(url)
				        return response

				    except requests.exceptions.RequestException as e:
				        print(e)

				def scrape_google(query):

					query = urllib.parse.quote_plus(query)
					response = get_source("https://www.google.com/search?q=" + query)

					links = list(response.html.absolute_links)
					google_domains = ('https://www.google.'
					)

					for url in links[:]:
						if url.startswith(google_domains):
							links.remove(url)

					return links

				link=scrape_google(col_name)
				return render(request,'engineering_search.html',{'row':row,'row1':row1,'a':a,'link':link,'row':row,
				'row1':row1,'a':a,'col_detail':col_detail,'row_1':row_1})
		elif caste == 'MBCDNC':
			a=Engineering_College_Detail.objects.filter(degree=degree,department=department,MBCDNC__lte=cutoff_mark)
			cursor=connection.cursor()
			sql1 = '''SELECT e.college_name,e.degree,e.department,e.MBCDNC from app_engineering_college_detail as e where e.degree='%s' 
			AND e.department='%s' AND e.MBCDNC <= '%s' ''' % (degree,department,cutoff_mark)
			post1=cursor.execute(sql1)
			row_1 = cursor.fetchall()
			cursor=connection.cursor()
			sql = '''SELECT e.college_name,e.degree,e.department from app_engineering_college_detail as e where e.degree='%s' 
			AND e.department='%s' AND e.MBCDNC='%s' ''' % (degree,department,cutoff_mark)
			post=cursor.execute(sql)
			col_detail = cursor.fetchone()
			if col_detail:
				col_name = col_detail[0]
				import requests
				import urllib
				import pandas as pd
				from requests_html import HTML
				from requests_html import HTMLSession
				def get_source(url):
				    """Return the source code for the provided URL. 

				    Args: 
				        url (string): URL of the page to scrape.

				    Returns:
				        response (object): HTTP response object from requests_html. 
				    """

				    try:
				        session = HTMLSession()
				        response = session.get(url)
				        return response

				    except requests.exceptions.RequestException as e:
				        print(e)

				def scrape_google(query):

					query = urllib.parse.quote_plus(query)
					response = get_source("https://www.google.com/search?q=" + query)

					links = list(response.html.absolute_links)
					google_domains = ('https://www.google.'
					)

					for url in links[:]:
						if url.startswith(google_domains):
							links.remove(url)

					return links

				link=scrape_google(col_name)
				return render(request,'engineering_search.html',{'link':link,'row':row,
				'row1':row1,'a':a,'col_detail':col_detail,'row_1':row_1})
		elif caste == 'MBA':
			a=Engineering_College_Detail.objects.filter(degree=degree,department=department,MBA__lte=cutoff_mark)
			cursor=connection.cursor()
			sql1 = '''SELECT e.college_name,e.degree,e.department,e.MBA from app_engineering_college_detail as e where e.degree='%s' 
			AND e.department='%s' AND e.MBA <= '%s' ''' % (degree,department,cutoff_mark)
			post1=cursor.execute(sql1)
			row_1 = cursor.fetchall()
			cursor=connection.cursor()
			sql = '''SELECT e.college_name,e.degree,e.department from app_engineering_college_detail as e where e.degree='%s' 
			AND e.department='%s' AND e.MBA='%s' ''' % (degree,department,cutoff_mark)
			post=cursor.execute(sql)
			col_detail = cursor.fetchone()
			if col_detail:
				col_name = col_detail[0]
				import requests
				import urllib
				import pandas as pd
				from requests_html import HTML
				from requests_html import HTMLSession
				def get_source(url):
				    """Return the source code for the provided URL. 

				    Args: 
				        url (string): URL of the page to scrape.

				    Returns:
				        response (object): HTTP response object from requests_html. 
				    """

				    try:
				        session = HTMLSession()
				        response = session.get(url)
				        return response

				    except requests.exceptions.RequestException as e:
				        print(e)

				def scrape_google(query):

					query = urllib.parse.quote_plus(query)
					response = get_source("https://www.google.com/search?q=" + query)

					links = list(response.html.absolute_links)
					google_domains = ('https://www.google.'
					)

					for url in links[:]:
						if url.startswith(google_domains):
							links.remove(url)

					return links

				link=scrape_google(col_name)
				return render(request,'engineering_search.html',{'link':link,'row':row,
				'row1':row1,'a':a,'col_detail':col_detail,'row_1':row_1})
		elif caste == 'ST':
			a=Engineering_College_Detail.objects.filter(degree=degree,department=department,ST__lte=cutoff_mark)
			cursor=connection.cursor()
			sql1 = '''SELECT e.college_name,e.degree,e.department,e.ST from app_engineering_college_detail as e where e.degree='%s' 
			AND e.department='%s' AND e.ST <= '%s' ''' % (degree,department,cutoff_mark)
			post1=cursor.execute(sql1)
			row_1 = cursor.fetchall()
			cursor=connection.cursor()
			sql = '''SELECT e.college_name,e.degree,e.department from app_engineering_college_detail as e where e.degree='%s' 
			AND e.department='%s' AND e.ST='%s' ''' % (degree,department,cutoff_mark)
			post=cursor.execute(sql)
			col_detail = cursor.fetchone()
			if col_detail:
				col_name = col_detail[0]
				import requests
				import urllib
				import pandas as pd
				from requests_html import HTML
				from requests_html import HTMLSession
				def get_source(url):
				    """Return the source code for the provided URL. 

				    Args: 
				        url (string): URL of the page to scrape.

				    Returns:
				        response (object): HTTP response object from requests_html. 
				    """

				    try:
				        session = HTMLSession()
				        response = session.get(url)
				        return response

				    except requests.exceptions.RequestException as e:
				        print(e)

				def scrape_google(query):

					query = urllib.parse.quote_plus(query)
					response = get_source("https://www.google.com/search?q=" + query)

					links = list(response.html.absolute_links)
					google_domains = ('https://www.google.'
					)

					for url in links[:]:
						if url.startswith(google_domains):
							links.remove(url)

					return links

				link=scrape_google(col_name)
				return render(request,'engineering_search.html',{'row':row,'row1':row1,'a':a,'link':link,'row':row,
				'row1':row1,'a':a,'col_detail':col_detail,'row_1':row_1})
		elif caste == 'SC':
			a=Engineering_College_Detail.objects.filter(degree=degree,department=department,SC__lte=cutoff_mark)
			cursor=connection.cursor()
			sql1 = '''SELECT e.college_name,e.degree,e.department,e.SC from app_engineering_college_detail as e where e.degree='%s' 
			AND e.department='%s' AND e.SC <= '%s' ''' % (degree,department,cutoff_mark)
			post1=cursor.execute(sql1)
			row_1 = cursor.fetchall()
			cursor=connection.cursor()
			sql = '''SELECT e.college_name,e.degree,e.department from app_engineering_college_detail as e where e.degree='%s' 
			AND e.department='%s' AND e.SC<='%s' ''' % (degree,department,cutoff_mark)
			post=cursor.execute(sql)
			col_detail = cursor.fetchone()
			if col_detail:
				col_name = col_detail[0]
				import requests
				import urllib
				import pandas as pd
				from requests_html import HTML
				from requests_html import HTMLSession
				def get_source(url):
				    """Return the source code for the provided URL. 

				    Args: 
				        url (string): URL of the page to scrape.

				    Returns:
				        response (object): HTTP response object from requests_html. 
				    """

				    try:
				        session = HTMLSession()
				        response = session.get(url)
				        return response

				    except requests.exceptions.RequestException as e:
				        print(e)

				def scrape_google(query):

					query = urllib.parse.quote_plus(query)
					response = get_source("https://www.google.com/search?q=" + query)

					links = list(response.html.absolute_links)
					google_domains = ('https://www.google.'
					)

					for url in links[:]:
						if url.startswith(google_domains):
							links.remove(url)

					return links

				link=scrape_google(col_name)
				return render(request,'engineering_search.html',{'link':link,'row':row,
				'row1':row1,'a':a,'col_detail':col_detail,'row_1':row_1})
		elif caste == 'SCA':
			a=Engineering_College_Detail.objects.filter(degree=degree,department=department,SCA__lte=cutoff_mark)
			cursor=connection.cursor()
			sql1 = '''SELECT e.college_name,e.degree,e.department,e.SCA from app_engineering_college_detail as e where e.degree='%s' 
			AND e.department='%s' AND e.SCA <= '%s' ''' % (degree,department,cutoff_mark)
			post1=cursor.execute(sql1)
			row_1 = cursor.fetchall()
			cursor=connection.cursor()
			sql = '''SELECT e.college_name,e.degree,e.department from app_engineering_college_detail as e where e.degree='%s' 
			AND e.department='%s' AND e.SCA='%s' ''' % (degree,department,cutoff_mark)
			post=cursor.execute(sql)
			col_detail = cursor.fetchone()
			if col_detail:
				col_name = col_detail[0]
				import requests
				import urllib
				import pandas as pd
				from requests_html import HTML
				from requests_html import HTMLSession
				def get_source(url):
				    """Return the source code for the provided URL. 

				    Args: 
				        url (string): URL of the page to scrape.

				    Returns:
				        response (object): HTTP response object from requests_html. 
				    """

				    try:
				        session = HTMLSession()
				        response = session.get(url)
				        return response

				    except requests.exceptions.RequestException as e:
				        print(e)

				def scrape_google(query):

					query = urllib.parse.quote_plus(query)
					response = get_source("https://www.google.com/search?q=" + query)

					links = list(response.html.absolute_links)
					google_domains = ('https://www.google.'
					)

					for url in links[:]:
						if url.startswith(google_domains):
							links.remove(url)

					return links

				link=scrape_google(col_name)
				return render(request,'engineering_search.html',{'link':link,'row':row,
				'row1':row1,'a':a,'col_detail':col_detail,'row_1':row_1})
	return render(request,'engineering_search.html',{'row':row,'row1':row1,'caste':caste})
def medical_search(request):
	user_id=request.session['user_id']
	cur=connection.cursor()
	user = '''SELECT s.caste from app_student_detail as s where s.id='%d' ''' %(int(user_id))
	sel = cur.execute(user)
	user_detail = cur.fetchone()
	caste = user_detail[0]
	cursor=connection.cursor()
	sql=''' SELECT e.degree from app_medical_college_detail as e GROUP BY e.degree'''
	post=cursor.execute(sql)
	row=cursor.fetchall()
	cursor1=connection.cursor()
	sql1=''' SELECT e.department from app_medical_college_detail as e GROUP BY e.department'''
	post1=cursor1.execute(sql1)
	row1=cursor1.fetchall()
	if request.method == 'POST':
		degree=request.POST.get('degree')
		department=request.POST.get('department')
		cutoff_mark =request.POST.get('cutoff_mark')
		if caste == 'OC':
			a=Medical_College_Detail.objects.filter(degree=degree,department=department,OC__lte=cutoff_mark)
			cursor=connection.cursor()
			sql = '''SELECT e.college_name,e.degree,e.department from app_medical_college_detail as e where e.degree='%s' 
			AND e.department='%s' AND e.OC='%s' ''' % (degree,department,cutoff_mark)
			post=cursor.execute(sql)
			col_detail = cursor.fetchone()
			if col_detail:
				col_name = col_detail[0]
				import requests
				import urllib
				import pandas as pd
				from requests_html import HTML
				from requests_html import HTMLSession
				def get_source(url):
				    """Return the source code for the provided URL. 

				    Args: 
				        url (string): URL of the page to scrape.

				    Returns:
				        response (object): HTTP response object from requests_html. 
				    """

				    try:
				        session = HTMLSession()
				        response = session.get(url)
				        return response

				    except requests.exceptions.RequestException as e:
				        print(e)

				def scrape_google(query):

					query = urllib.parse.quote_plus(query)
					response = get_source("https://www.google.com/search?q=" + query)

					links = list(response.html.absolute_links)
					google_domains = ('https://www.google.'
					)

					for url in links[:]:
						if url.startswith(google_domains):
							links.remove(url)

					return links

				link=scrape_google(col_name)
				

				return render(request,'medical_search.html',{'col_detail':col_detail,'row':row,'row1':row1,'a':a,'link':link})
		elif caste == 'BC':
			a=Medical_College_Detail.objects.filter(degree=degree,department=department,BC__lte=cutoff_mark)
			cursor=connection.cursor()
			sql1 = '''SELECT e.college_name,e.degree,e.department,e.BC from app_medical_college_detail as e where e.degree='%s' 
			AND e.department='%s' AND e.BC <= '%s' ''' % (degree,department,cutoff_mark)
			post1=cursor.execute(sql1)
			row_1 = cursor.fetchall()
			cursor=connection.cursor()
			sql = '''SELECT e.college_name,e.degree,e.department,e.BC from app_medical_college_detail as e where e.degree='%s' 
			AND e.department='%s' AND e.BC <= '%s' ''' % (degree,department,cutoff_mark)
			post=cursor.execute(sql)
			col_detail = cursor.fetchone()
			if col_detail:
				col_name = col_detail[0]
				import requests
				import urllib
				import pandas as pd
				from requests_html import HTML
				from requests_html import HTMLSession
				def get_source(url):
				    """Return the source code for the provided URL. 

				    Args: 
				        url (string): URL of the page to scrape.

				    Returns:
				        response (object): HTTP response object from requests_html. 
				    """

				    try:
				        session = HTMLSession()
				        response = session.get(url)
				        return response

				    except requests.exceptions.RequestException as e:
				        print(e)

				def scrape_google(query):

					query = urllib.parse.quote_plus(query)
					response = get_source("https://www.google.com/search?q=" + query)

					links = list(response.html.absolute_links)
					google_domains = ('https://www.google.'
					)

					for url in links[:]:
						if url.startswith(google_domains):
							links.remove(url)

					return links

				link=scrape_google(col_name)
				return render(request,'medical_search.html',{'link':link,'row':row,
				'row1':row1,'a':a,'col_detail':col_detail,'row_1':row_1,'caste':caste})
		elif caste == 'BCM':
			a=Medical_College_Detail.objects.filter(degree=degree,department=department,BCM__lte=cutoff_mark)
			cursor=connection.cursor()
			sql1 = '''SELECT e.college_name,e.degree,e.department,e.BCM from app_medical_college_detail as e where e.degree='%s' 
			AND e.department='%s' AND e.BCM <= '%s' ''' % (degree,department,cutoff_mark)
			post1=cursor.execute(sql1)
			row_1 = cursor.fetchall()
			cursor=connection.cursor()
			sql = '''SELECT e.college_name,e.degree,e.department,e.BCM from app_medical_college_detail as e where e.degree='%s' 
			AND e.department='%s' AND e.BCM='%s' ''' % (degree,department,cutoff_mark)
			post=cursor.execute(sql)
			col_detail = cursor.fetchone()
			if col_detail:
				col_name = col_detail[0]
				import requests
				import urllib
				import pandas as pd
				from requests_html import HTML
				from requests_html import HTMLSession
				def get_source(url):
				    """Return the source code for the provided URL. 

				    Args: 
				        url (string): URL of the page to scrape.

				    Returns:
				        response (object): HTTP response object from requests_html. 
				    """

				    try:
				        session = HTMLSession()
				        response = session.get(url)
				        return response

				    except requests.exceptions.RequestException as e:
				        print(e)

				def scrape_google(query):

					query = urllib.parse.quote_plus(query)
					response = get_source("https://www.google.com/search?q=" + query)

					links = list(response.html.absolute_links)
					google_domains = ('https://www.google.'
					)

					for url in links[:]:
						if url.startswith(google_domains):
							links.remove(url)

					return links

				link=scrape_google(col_name)
				return render(request,'medical_search.html',{'link':link,'row':row,
				'row1':row1,'a':a,'col_detail':col_detail,'row_1':row_1})
		elif caste == 'MBCV':
			a=Medical_College_Detail.objects.filter(degree=degree,department=department,MBCV__lte=cutoff_mark)
			cursor=connection.cursor()
			sql1 = '''SELECT e.college_name,e.degree,e.department,e.MBCV from app_medical_college_detail as e where e.degree='%s' 
			AND e.department='%s' AND e.MBCV <= '%s' ''' % (degree,department,cutoff_mark)
			post1=cursor.execute(sql1)
			row_1 = cursor.fetchall()
			cursor=connection.cursor()
			sql = '''SELECT e.college_name,e.degree,e.department from app_medical_college_detail as e where e.degree='%s' 
			AND e.department='%s' AND e.MBCV='%s' ''' % (degree,department,cutoff_mark)
			post=cursor.execute(sql)
			col_detail = cursor.fetchone()
			if col_detail:
				col_name = col_detail[0]
				import requests
				import urllib
				import pandas as pd
				from requests_html import HTML
				from requests_html import HTMLSession
				def get_source(url):
				    """Return the source code for the provided URL. 

				    Args: 
				        url (string): URL of the page to scrape.

				    Returns:
				        response (object): HTTP response object from requests_html. 
				    """

				    try:
				        session = HTMLSession()
				        response = session.get(url)
				        return response

				    except requests.exceptions.RequestException as e:
				        print(e)

				def scrape_google(query):

					query = urllib.parse.quote_plus(query)
					response = get_source("https://www.google.com/search?q=" + query)

					links = list(response.html.absolute_links)
					google_domains = ('https://www.google.'
					)

					for url in links[:]:
						if url.startswith(google_domains):
							links.remove(url)

					return links

				link=scrape_google(col_name)
				return render(request,'medical_search.html',{'row':row,'row1':row1,'a':a,'link':link,'row':row,
				'row1':row1,'a':a,'col_detail':col_detail,'row_1':row_1})
		elif caste == 'MBCDNC':
			a=Medical_College_Detail.objects.filter(degree=degree,department=department,MBCDNC__lte=cutoff_mark)
			cursor=connection.cursor()
			sql1 = '''SELECT e.college_name,e.degree,e.department,e.MBCDNC from app_medical_college_detail as e where e.degree='%s' 
			AND e.department='%s' AND e.MBCDNC <= '%s' ''' % (degree,department,cutoff_mark)
			post1=cursor.execute(sql1)
			row_1 = cursor.fetchall()
			cursor=connection.cursor()
			sql = '''SELECT e.college_name,e.degree,e.department from app_medical_college_detail as e where e.degree='%s' 
			AND e.department='%s' AND e.MBCDNC='%s' ''' % (degree,department,cutoff_mark)
			post=cursor.execute(sql)
			col_detail = cursor.fetchone()
			if col_detail:
				col_name = col_detail[0]
				import requests
				import urllib
				import pandas as pd
				from requests_html import HTML
				from requests_html import HTMLSession
				def get_source(url):
				    """Return the source code for the provided URL. 

				    Args: 
				        url (string): URL of the page to scrape.

				    Returns:
				        response (object): HTTP response object from requests_html. 
				    """

				    try:
				        session = HTMLSession()
				        response = session.get(url)
				        return response

				    except requests.exceptions.RequestException as e:
				        print(e)

				def scrape_google(query):

					query = urllib.parse.quote_plus(query)
					response = get_source("https://www.google.com/search?q=" + query)

					links = list(response.html.absolute_links)
					google_domains = ('https://www.google.'
					)

					for url in links[:]:
						if url.startswith(google_domains):
							links.remove(url)

					return links

				link=scrape_google(col_name)
				return render(request,'medical_search.html',{'link':link,'row':row,
				'row1':row1,'a':a,'col_detail':col_detail,'row_1':row_1})
		elif caste == 'MBA':
			a=Medical_College_Detail.objects.filter(degree=degree,department=department,MBA__lte=cutoff_mark)
			cursor=connection.cursor()
			sql1 = '''SELECT e.college_name,e.degree,e.department,e.MBA from app_medical_college_detail as e where e.degree='%s' 
			AND e.department='%s' AND e.MBA <= '%s' ''' % (degree,department,cutoff_mark)
			post1=cursor.execute(sql1)
			row_1 = cursor.fetchall()
			cursor=connection.cursor()
			sql = '''SELECT e.college_name,e.degree,e.department from app_medical_college_detail as e where e.degree='%s' 
			AND e.department='%s' AND e.MBA='%s' ''' % (degree,department,cutoff_mark)
			post=cursor.execute(sql)
			col_detail = cursor.fetchone()
			if col_detail:
				col_name = col_detail[0]
				import requests
				import urllib
				import pandas as pd
				from requests_html import HTML
				from requests_html import HTMLSession
				def get_source(url):
				    """Return the source code for the provided URL. 

				    Args: 
				        url (string): URL of the page to scrape.

				    Returns:
				        response (object): HTTP response object from requests_html. 
				    """

				    try:
				        session = HTMLSession()
				        response = session.get(url)
				        return response

				    except requests.exceptions.RequestException as e:
				        print(e)

				def scrape_google(query):

					query = urllib.parse.quote_plus(query)
					response = get_source("https://www.google.com/search?q=" + query)

					links = list(response.html.absolute_links)
					google_domains = ('https://www.google.'
					)

					for url in links[:]:
						if url.startswith(google_domains):
							links.remove(url)

					return links

				link=scrape_google(col_name)
				return render(request,'medical_search.html',{'link':link,'row':row,
				'row1':row1,'a':a,'col_detail':col_detail,'row_1':row_1})
		elif caste == 'ST':
			a=Medical_College_Detail.objects.filter(degree=degree,department=department,ST__lte=cutoff_mark)
			cursor=connection.cursor()
			sql1 = '''SELECT e.college_name,e.degree,e.department,e.ST from app_medical_college_detail as e where e.degree='%s' 
			AND e.department='%s' AND e.ST <= '%s' ''' % (degree,department,cutoff_mark)
			post1=cursor.execute(sql1)
			row_1 = cursor.fetchall()
			cursor=connection.cursor()
			sql = '''SELECT e.college_name,e.degree,e.department from app_medical_college_detail as e where e.degree='%s' 
			AND e.department='%s' AND e.ST='%s' ''' % (degree,department,cutoff_mark)
			post=cursor.execute(sql)
			col_detail = cursor.fetchone()
			if col_detail:
				col_name = col_detail[0]
				import requests
				import urllib
				import pandas as pd
				from requests_html import HTML
				from requests_html import HTMLSession
				def get_source(url):
				    """Return the source code for the provided URL. 

				    Args: 
				        url (string): URL of the page to scrape.

				    Returns:
				        response (object): HTTP response object from requests_html. 
				    """

				    try:
				        session = HTMLSession()
				        response = session.get(url)
				        return response

				    except requests.exceptions.RequestException as e:
				        print(e)

				def scrape_google(query):

					query = urllib.parse.quote_plus(query)
					response = get_source("https://www.google.com/search?q=" + query)

					links = list(response.html.absolute_links)
					google_domains = ('https://www.google.'
					)

					for url in links[:]:
						if url.startswith(google_domains):
							links.remove(url)

					return links

				link=scrape_google(col_name)
				return render(request,'medical_search.html',{'row':row,'row1':row1,'a':a,'link':link,'row':row,
				'row1':row1,'a':a,'col_detail':col_detail,'row_1':row_1})
		elif caste == 'SC':
			a=Medical_College_Detail.objects.filter(degree=degree,department=department,SC__lte=cutoff_mark)
			cursor=connection.cursor()
			sql1 = '''SELECT e.college_name,e.degree,e.department,e.SC from app_medical_college_detail as e where e.degree='%s' 
			AND e.department='%s' AND e.SC <= '%s' ''' % (degree,department,cutoff_mark)
			post1=cursor.execute(sql1)
			row_1 = cursor.fetchall()
			cursor=connection.cursor()
			sql = '''SELECT e.college_name,e.degree,e.department from app_medical_college_detail as e where e.degree='%s' 
			AND e.department='%s' AND e.SC<='%s' ''' % (degree,department,cutoff_mark)
			post=cursor.execute(sql)
			col_detail = cursor.fetchone()
			if col_detail:
				col_name = col_detail[0]
				import requests
				import urllib
				import pandas as pd
				from requests_html import HTML
				from requests_html import HTMLSession
				def get_source(url):
				    """Return the source code for the provided URL. 

				    Args: 
				        url (string): URL of the page to scrape.

				    Returns:
				        response (object): HTTP response object from requests_html. 
				    """

				    try:
				        session = HTMLSession()
				        response = session.get(url)
				        return response

				    except requests.exceptions.RequestException as e:
				        print(e)

				def scrape_google(query):

					query = urllib.parse.quote_plus(query)
					response = get_source("https://www.google.com/search?q=" + query)

					links = list(response.html.absolute_links)
					google_domains = ('https://www.google.'
					)

					for url in links[:]:
						if url.startswith(google_domains):
							links.remove(url)

					return links

				link=scrape_google(col_name)
				return render(request,'medical_search.html',{'link':link,'row':row,
				'row1':row1,'a':a,'col_detail':col_detail,'row_1':row_1})
		elif caste == 'SCA':
			a=Medical_College_Detail.objects.filter(degree=degree,department=department,SCA__lte=cutoff_mark)
			cursor=connection.cursor()
			sql1 = '''SELECT e.college_name,e.degree,e.department,e.SCA from app_medical_college_detail as e where e.degree='%s' 
			AND e.department='%s' AND e.SCA <= '%s' ''' % (degree,department,cutoff_mark)
			post1=cursor.execute(sql1)
			row_1 = cursor.fetchall()
			cursor=connection.cursor()
			sql = '''SELECT e.college_name,e.degree,e.department from app_medical_college_detail as e where e.degree='%s' 
			AND e.department='%s' AND e.SCA='%s' ''' % (degree,department,cutoff_mark)
			post=cursor.execute(sql)
			col_detail = cursor.fetchone()
			if col_detail:
				col_name = col_detail[0]
				import requests
				import urllib
				import pandas as pd
				from requests_html import HTML
				from requests_html import HTMLSession
				def get_source(url):
				    """Return the source code for the provided URL. 

				    Args: 
				        url (string): URL of the page to scrape.

				    Returns:
				        response (object): HTTP response object from requests_html. 
				    """

				    try:
				        session = HTMLSession()
				        response = session.get(url)
				        return response

				    except requests.exceptions.RequestException as e:
				        print(e)

				def scrape_google(query):

					query = urllib.parse.quote_plus(query)
					response = get_source("https://www.google.com/search?q=" + query)

					links = list(response.html.absolute_links)
					google_domains = ('https://www.google.'
					)

					for url in links[:]:
						if url.startswith(google_domains):
							links.remove(url)

					return links

				link=scrape_google(col_name)
				return render(request,'medical_search.html',{'link':link,'row':row,
				'row1':row1,'a':a,'col_detail':col_detail,'row_1':row_1})
	return render(request,'medical_search.html',{'row':row,'row1':row1,'caste':caste})
def agri_colg_search(request):
	user_id=request.session['user_id']
	cur=connection.cursor()
	user = '''SELECT s.caste from app_student_detail as s where s.id='%d' ''' %(int(user_id))
	sel = cur.execute(user)
	user_detail = cur.fetchone()
	caste = user_detail[0]
	cursor=connection.cursor()
	sql=''' SELECT e.degree from app_agri_college_detail as e GROUP BY e.degree'''
	post=cursor.execute(sql)
	row=cursor.fetchall()
	cursor1=connection.cursor()
	sql1=''' SELECT e.department from app_agri_college_detail as e GROUP BY e.department'''
	post1=cursor1.execute(sql1)
	row1=cursor1.fetchall()
	if request.method == 'POST':
		degree=request.POST.get('degree')
		department=request.POST.get('department')
		cutoff_mark =request.POST.get('cutoff_mark')
		if caste == 'OC':
			a=Agri_College_Detail.objects.filter(degree=degree,department=department,OC__lte=cutoff_mark)
			cursor=connection.cursor()
			sql = '''SELECT e.college_name,e.degree,e.department from app_agri_college_detail as e where e.degree='%s' 
			AND e.department='%s' AND e.OC='%s' ''' % (degree,department,cutoff_mark)
			post=cursor.execute(sql)
			col_detail = cursor.fetchone()
			if col_detail:
				col_name = col_detail[0]
				import requests
				import urllib
				import pandas as pd
				from requests_html import HTML
				from requests_html import HTMLSession
				def get_source(url):
				    """Return the source code for the provided URL. 

				    Args: 
				        url (string): URL of the page to scrape.

				    Returns:
				        response (object): HTTP response object from requests_html. 
				    """

				    try:
				        session = HTMLSession()
				        response = session.get(url)
				        return response

				    except requests.exceptions.RequestException as e:
				        print(e)

				def scrape_google(query):

					query = urllib.parse.quote_plus(query)
					response = get_source("https://www.google.com/search?q=" + query)

					links = list(response.html.absolute_links)
					google_domains = ('https://www.google.'
					)

					for url in links[:]:
						if url.startswith(google_domains):
							links.remove(url)

					return links

				link=scrape_google(col_name)
				

				return render(request,'agri_colg_search.html',{'col_detail':col_detail,'row':row,'row1':row1,'a':a,'link':link})
		elif caste == 'BC':
			a=Agri_College_Detail.objects.filter(degree=degree,department=department,BC__lte=cutoff_mark)
			cursor=connection.cursor()
			sql1 = '''SELECT e.college_name,e.degree,e.department,e.BC from app_agri_college_detail as e where e.degree='%s' 
			AND e.department='%s' AND e.BC <= '%s' ''' % (degree,department,cutoff_mark)
			post1=cursor.execute(sql1)
			row_1 = cursor.fetchall()
			cursor=connection.cursor()
			sql = '''SELECT e.college_name,e.degree,e.department,e.BC from app_agri_college_detail as e where e.degree='%s' 
			AND e.department='%s' AND e.BC <= '%s' ''' % (degree,department,cutoff_mark)
			post=cursor.execute(sql)
			col_detail = cursor.fetchone()
			if col_detail:
				col_name = col_detail[0]
				import requests
				import urllib
				import pandas as pd
				from requests_html import HTML
				from requests_html import HTMLSession
				def get_source(url):
				    """Return the source code for the provided URL. 

				    Args: 
				        url (string): URL of the page to scrape.

				    Returns:
				        response (object): HTTP response object from requests_html. 
				    """

				    try:
				        session = HTMLSession()
				        response = session.get(url)
				        return response

				    except requests.exceptions.RequestException as e:
				        print(e)

				def scrape_google(query):

					query = urllib.parse.quote_plus(query)
					response = get_source("https://www.google.com/search?q=" + query)

					links = list(response.html.absolute_links)
					google_domains = ('https://www.google.'
					)

					for url in links[:]:
						if url.startswith(google_domains):
							links.remove(url)

					return links

				link=scrape_google(col_name)
				return render(request,'agri_colg_search.html',{'link':link,'row':row,
				'row1':row1,'a':a,'col_detail':col_detail,'row_1':row_1})
		elif caste == 'BCM':
			a=Agri_College_Detail.objects.filter(degree=degree,department=department,BCM__lte=cutoff_mark)
			cursor=connection.cursor()
			sql1 = '''SELECT e.college_name,e.degree,e.department,e.BCM from app_agri_college_detail as e where e.degree='%s' 
			AND e.department='%s' AND e.BCM <= '%s' ''' % (degree,department,cutoff_mark)
			post1=cursor.execute(sql1)
			row_1 = cursor.fetchall()
			cursor=connection.cursor()
			sql = '''SELECT e.college_name,e.degree,e.department,e.BCM from app_agri_college_detail as e where e.degree='%s' 
			AND e.department='%s' AND e.BCM='%s' ''' % (degree,department,cutoff_mark)
			post=cursor.execute(sql)
			col_detail = cursor.fetchone()
			if col_detail:
				col_name = col_detail[0]
				import requests
				import urllib
				import pandas as pd
				from requests_html import HTML
				from requests_html import HTMLSession
				def get_source(url):
				    """Return the source code for the provided URL. 

				    Args: 
				        url (string): URL of the page to scrape.

				    Returns:
				        response (object): HTTP response object from requests_html. 
				    """

				    try:
				        session = HTMLSession()
				        response = session.get(url)
				        return response

				    except requests.exceptions.RequestException as e:
				        print(e)

				def scrape_google(query):

					query = urllib.parse.quote_plus(query)
					response = get_source("https://www.google.com/search?q=" + query)

					links = list(response.html.absolute_links)
					google_domains = ('https://www.google.'
					)

					for url in links[:]:
						if url.startswith(google_domains):
							links.remove(url)

					return links

				link=scrape_google(col_name)
				return render(request,'agri_colg_search.html',{'link':link,'row':row,
				'row1':row1,'a':a,'col_detail':col_detail,'row_1':row_1})
		elif caste == 'MBCV':
			a=Agri_College_Detail.objects.filter(degree=degree,department=department,MBCV__lte=cutoff_mark)
			cursor=connection.cursor()
			sql1 = '''SELECT e.college_name,e.degree,e.department,e.MBCV from app_agri_college_detail as e where e.degree='%s' 
			AND e.department='%s' AND e.MBCV <= '%s' ''' % (degree,department,cutoff_mark)
			post1=cursor.execute(sql1)
			row_1 = cursor.fetchall()
			cursor=connection.cursor()
			sql = '''SELECT e.college_name,e.degree,e.department from app_agri_college_detail as e where e.degree='%s' 
			AND e.department='%s' AND e.MBCV='%s' ''' % (degree,department,cutoff_mark)
			post=cursor.execute(sql)
			col_detail = cursor.fetchone()
			if col_detail:
				col_name = col_detail[0]
				import requests
				import urllib
				import pandas as pd
				from requests_html import HTML
				from requests_html import HTMLSession
				def get_source(url):
				    """Return the source code for the provided URL. 

				    Args: 
				        url (string): URL of the page to scrape.

				    Returns:
				        response (object): HTTP response object from requests_html. 
				    """

				    try:
				        session = HTMLSession()
				        response = session.get(url)
				        return response

				    except requests.exceptions.RequestException as e:
				        print(e)

				def scrape_google(query):

					query = urllib.parse.quote_plus(query)
					response = get_source("https://www.google.com/search?q=" + query)

					links = list(response.html.absolute_links)
					google_domains = ('https://www.google.'
					)

					for url in links[:]:
						if url.startswith(google_domains):
							links.remove(url)

					return links

				link=scrape_google(col_name)
				return render(request,'agri_colg_search.html',{'row':row,'row1':row1,'a':a,'link':link,'row':row,
				'row1':row1,'a':a,'col_detail':col_detail,'row_1':row_1})
		elif caste == 'MBCDNC':
			a=Agri_College_Detail.objects.filter(degree=degree,department=department,MBCDNC__lte=cutoff_mark)
			cursor=connection.cursor()
			sql1 = '''SELECT e.college_name,e.degree,e.department,e.MBCDNC from app_agri_college_detail as e where e.degree='%s' 
			AND e.department='%s' AND e.MBCDNC <= '%s' ''' % (degree,department,cutoff_mark)
			post1=cursor.execute(sql1)
			row_1 = cursor.fetchall()
			cursor=connection.cursor()
			sql = '''SELECT e.college_name,e.degree,e.department from app_agri_college_detail as e where e.degree='%s' 
			AND e.department='%s' AND e.MBCDNC='%s' ''' % (degree,department,cutoff_mark)
			post=cursor.execute(sql)
			col_detail = cursor.fetchone()
			if col_detail:
				col_name = col_detail[0]
				import requests
				import urllib
				import pandas as pd
				from requests_html import HTML
				from requests_html import HTMLSession
				def get_source(url):
				    """Return the source code for the provided URL. 

				    Args: 
				        url (string): URL of the page to scrape.

				    Returns:
				        response (object): HTTP response object from requests_html. 
				    """

				    try:
				        session = HTMLSession()
				        response = session.get(url)
				        return response

				    except requests.exceptions.RequestException as e:
				        print(e)

				def scrape_google(query):

					query = urllib.parse.quote_plus(query)
					response = get_source("https://www.google.com/search?q=" + query)

					links = list(response.html.absolute_links)
					google_domains = ('https://www.google.'
					)

					for url in links[:]:
						if url.startswith(google_domains):
							links.remove(url)

					return links

				link=scrape_google(col_name)
				return render(request,'agri_colg_search.html',{'link':link,'row':row,
				'row1':row1,'a':a,'col_detail':col_detail,'row_1':row_1})
		elif caste == 'MBA':
			a=Agri_College_Detail.objects.filter(degree=degree,department=department,MBA__lte=cutoff_mark)
			cursor=connection.cursor()
			sql1 = '''SELECT e.college_name,e.degree,e.department,e.MBA from app_agri_college_detail as e where e.degree='%s' 
			AND e.department='%s' AND e.MBA <= '%s' ''' % (degree,department,cutoff_mark)
			post1=cursor.execute(sql1)
			row_1 = cursor.fetchall()
			cursor=connection.cursor()
			sql = '''SELECT e.college_name,e.degree,e.department from app_agri_college_detail as e where e.degree='%s' 
			AND e.department='%s' AND e.MBA='%s' ''' % (degree,department,cutoff_mark)
			post=cursor.execute(sql)
			col_detail = cursor.fetchone()
			if col_detail:
				col_name = col_detail[0]
				import requests
				import urllib
				import pandas as pd
				from requests_html import HTML
				from requests_html import HTMLSession
				def get_source(url):
				    """Return the source code for the provided URL. 

				    Args: 
				        url (string): URL of the page to scrape.

				    Returns:
				        response (object): HTTP response object from requests_html. 
				    """

				    try:
				        session = HTMLSession()
				        response = session.get(url)
				        return response

				    except requests.exceptions.RequestException as e:
				        print(e)

				def scrape_google(query):

					query = urllib.parse.quote_plus(query)
					response = get_source("https://www.google.com/search?q=" + query)

					links = list(response.html.absolute_links)
					google_domains = ('https://www.google.'
					)

					for url in links[:]:
						if url.startswith(google_domains):
							links.remove(url)

					return links

				link=scrape_google(col_name)
				return render(request,'agri_colg_search.html',{'link':link,'row':row,
				'row1':row1,'a':a,'col_detail':col_detail,'row_1':row_1})
		elif caste == 'ST':
			a=Agri_College_Detail.objects.filter(degree=degree,department=department,ST__lte=cutoff_mark)
			cursor=connection.cursor()
			sql1 = '''SELECT e.college_name,e.degree,e.department,e.ST from app_agri_college_detail as e where e.degree='%s' 
			AND e.department='%s' AND e.ST <= '%s' ''' % (degree,department,cutoff_mark)
			post1=cursor.execute(sql1)
			row_1 = cursor.fetchall()
			cursor=connection.cursor()
			sql = '''SELECT e.college_name,e.degree,e.department from app_agri_college_detail as e where e.degree='%s' 
			AND e.department='%s' AND e.ST='%s' ''' % (degree,department,cutoff_mark)
			post=cursor.execute(sql)
			col_detail = cursor.fetchone()
			if col_detail:
				col_name = col_detail[0]
				import requests
				import urllib
				import pandas as pd
				from requests_html import HTML
				from requests_html import HTMLSession
				def get_source(url):
				    """Return the source code for the provided URL. 

				    Args: 
				        url (string): URL of the page to scrape.

				    Returns:
				        response (object): HTTP response object from requests_html. 
				    """

				    try:
				        session = HTMLSession()
				        response = session.get(url)
				        return response

				    except requests.exceptions.RequestException as e:
				        print(e)

				def scrape_google(query):

					query = urllib.parse.quote_plus(query)
					response = get_source("https://www.google.com/search?q=" + query)

					links = list(response.html.absolute_links)
					google_domains = ('https://www.google.'
					)

					for url in links[:]:
						if url.startswith(google_domains):
							links.remove(url)

					return links

				link=scrape_google(col_name)
				return render(request,'agri_colg_search.html',{'row':row,'row1':row1,'a':a,'link':link,'row':row,
				'row1':row1,'a':a,'col_detail':col_detail,'row_1':row_1})
		elif caste == 'SC':
			a=Agri_College_Detail.objects.filter(degree=degree,department=department,SC__lte=cutoff_mark)
			cursor=connection.cursor()
			sql1 = '''SELECT e.college_name,e.degree,e.department,e.SC from app_agri_college_detail as e where e.degree='%s' 
			AND e.department='%s' AND e.SC <= '%s' ''' % (degree,department,cutoff_mark)
			post1=cursor.execute(sql1)
			row_1 = cursor.fetchall()
			cursor=connection.cursor()
			sql = '''SELECT e.college_name,e.degree,e.department from app_agri_college_detail as e where e.degree='%s' 
			AND e.department='%s' AND e.SC<='%s' ''' % (degree,department,cutoff_mark)
			post=cursor.execute(sql)
			col_detail = cursor.fetchone()
			if col_detail:
				col_name = col_detail[0]
				import requests
				import urllib
				import pandas as pd
				from requests_html import HTML
				from requests_html import HTMLSession
				def get_source(url):
				    """Return the source code for the provided URL. 

				    Args: 
				        url (string): URL of the page to scrape.

				    Returns:
				        response (object): HTTP response object from requests_html. 
				    """

				    try:
				        session = HTMLSession()
				        response = session.get(url)
				        return response

				    except requests.exceptions.RequestException as e:
				        print(e)

				def scrape_google(query):

					query = urllib.parse.quote_plus(query)
					response = get_source("https://www.google.com/search?q=" + query)

					links = list(response.html.absolute_links)
					google_domains = ('https://www.google.'
					)

					for url in links[:]:
						if url.startswith(google_domains):
							links.remove(url)

					return links

				link=scrape_google(col_name)
				return render(request,'agri_colg_search.html',{'link':link,'row':row,
				'row1':row1,'a':a,'col_detail':col_detail,'row_1':row_1})
		elif caste == 'SCA':
			a=Agri_College_Detail.objects.filter(degree=degree,department=department,SCA__lte=cutoff_mark)
			cursor=connection.cursor()
			sql1 = '''SELECT e.college_name,e.degree,e.department,e.SCA from app_agri_college_detail as e where e.degree='%s' 
			AND e.department='%s' AND e.SCA <= '%s' ''' % (degree,department,cutoff_mark)
			post1=cursor.execute(sql1)
			row_1 = cursor.fetchall()
			cursor=connection.cursor()
			sql = '''SELECT e.college_name,e.degree,e.department from app_agri_college_detail as e where e.degree='%s' 
			AND e.department='%s' AND e.SCA='%s' ''' % (degree,department,cutoff_mark)
			post=cursor.execute(sql)
			col_detail = cursor.fetchone()
			if col_detail:
				col_name = col_detail[0]
				import requests
				import urllib
				import pandas as pd
				from requests_html import HTML
				from requests_html import HTMLSession
				def get_source(url):
				    """Return the source code for the provided URL. 

				    Args: 
				        url (string): URL of the page to scrape.

				    Returns:
				        response (object): HTTP response object from requests_html. 
				    """

				    try:
				        session = HTMLSession()
				        response = session.get(url)
				        return response

				    except requests.exceptions.RequestException as e:
				        print(e)

				def scrape_google(query):

					query = urllib.parse.quote_plus(query)
					response = get_source("https://www.google.com/search?q=" + query)

					links = list(response.html.absolute_links)
					google_domains = ('https://www.google.'
					)

					for url in links[:]:
						if url.startswith(google_domains):
							links.remove(url)

					return links

				link=scrape_google(col_name)
				return render(request,'agri_colg_search.html',{'link':link,'row':row,
				'row1':row1,'a':a,'col_detail':col_detail,'row_1':row_1})
	return render(request,'agri_colg_search.html',{'row':row,'row1':row1,'caste':caste})